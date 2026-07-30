from __future__ import annotations

import math
import os
import re
import shutil
import subprocess
import tempfile
import threading
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional
from xml.etree import ElementTree as ET

from rich.text import Text
from textual import events, on, work
from textual.app import ComposeResult
from textual.containers import Horizontal, ScrollableContainer, Vertical
from textual.message import Message
from textual.timer import Timer
from textual.widgets import Button, Static

from .header import rich_preview_title

IMAGE_EXTENSIONS = {
    ".bmp",
    ".dib",
    ".gif",
    ".heic",
    ".heif",
    ".ico",
    ".jfif",
    ".jpeg",
    ".jpg",
    ".png",
    ".tif",
    ".tiff",
    ".webp",
}
PDF_EXTENSIONS = {".pdf"}
EXCEL_EXTENSIONS = {
    ".xls",
    ".xlsm",
    ".xlsx",
    ".xltm",
    ".xltx",
}
PREVIEW_EXTENSIONS = IMAGE_EXTENSIONS | PDF_EXTENSIONS | EXCEL_EXTENSIONS
MAX_EXCEL_ROWS = 35
MAX_EXCEL_COLUMNS = 12
MAX_PREVIEW_SOURCE_PIXELS = 48_000_000


@dataclass(frozen=True)
class DocumentPreview:
    """Rendered document preview and metadata for the right-side panel."""

    content: Text
    original_size: tuple[int, int]
    rendered_size: tuple[int, int]
    cell_size: tuple[int, int]
    kind: str
    detail: str
    scale: float
    fit_scale: float
    crop_box: tuple[int, int, int, int]
    native_pixels: bool


@dataclass
class DocumentSource:
    """Prepared preview image retained while one file remains selected."""

    image: object
    kind: str
    detail: str
    lock: threading.RLock = field(default_factory=threading.RLock)

    @property
    def size(self) -> tuple[int, int]:
        with self.lock:
            return self.image.size

    def copy_image(self):
        with self.lock:
            return self.image.copy()

    def close(self) -> None:
        with self.lock:
            try:
                self.image.close()
            except Exception:
                pass


def can_preview(path: Optional[Path]) -> bool:
    """Check a selected entry by extension without touching a slow drive."""
    return bool(
        path
        and path.suffix.lower() in PREVIEW_EXTENSIONS
    )


def _render_source_view(
    source: DocumentSource,
    *,
    max_columns: int,
    max_rows: int,
    zoom_factor: float = 1.0,
    native_pixels: bool = False,
    focus: tuple[float, float] = (0.5, 0.5),
) -> DocumentPreview:
    """Render a fitted or cropped view without allocating the full zoomed image."""
    from PIL import Image

    image = source.copy_image()
    original_size = image.size
    image_width, image_height = original_size
    max_columns = max(1, int(max_columns))
    max_rows = max(1, int(max_rows))
    fit_scale = min(
        1.0,
        max_columns / max(1, image_width),
        (max_rows * 2) / max(1, image_height),
    )
    scale = (
        1.0
        if native_pixels
        else min(1.0, fit_scale * max(1.0, float(zoom_factor)))
    )

    visible_source_width = min(
        image_width,
        max(1, int(math.floor(max_columns / max(scale, 1e-9)))),
    )
    visible_source_height = min(
        image_height,
        max(1, int(math.floor((max_rows * 2) / max(scale, 1e-9)))),
    )
    focus_x = max(0.0, min(1.0, float(focus[0])))
    focus_y = max(0.0, min(1.0, float(focus[1])))
    center_x = int(round(focus_x * image_width))
    center_y = int(round(focus_y * image_height))
    left = max(
        0,
        min(
            image_width - visible_source_width,
            center_x - visible_source_width // 2,
        ),
    )
    top = max(
        0,
        min(
            image_height - visible_source_height,
            center_y - visible_source_height // 2,
        ),
    )
    right = min(image_width, left + visible_source_width)
    bottom = min(image_height, top + visible_source_height)
    crop_box = (left, top, right, bottom)
    cropped = image.crop(crop_box).convert("RGBA")
    image.close()

    width = max(1, min(max_columns, int(round(cropped.width * scale))))
    height = max(1, min(max_rows * 2, int(round(cropped.height * scale))))
    if scale == 1.0 and cropped.size == (width, height):
        resized = cropped
    else:
        resampling = getattr(Image, "Resampling", Image)
        resized = cropped.resize(
            (width, height),
            resample=resampling.LANCZOS,
        )
        cropped.close()

    background = Image.new("RGB", (width, height), (18, 18, 18))
    background.paste(resized, mask=resized.getchannel("A"))
    resized.close()
    pixels = background.load()

    output = Text(no_wrap=True)
    for y in range(0, height, 2):
        bottom_y = min(y + 1, height - 1)
        for x in range(width):
            top = pixels[x, y]
            bottom = pixels[x, bottom_y]
            output.append(
                "\u2580",
                style=(
                    f"rgb({top[0]},{top[1]},{top[2]}) "
                    f"on rgb({bottom[0]},{bottom[1]},{bottom[2]})"
                ),
            )
        if y + 2 < height:
            output.append("\n")

    preview = DocumentPreview(
        content=output,
        original_size=original_size,
        rendered_size=(width, height),
        cell_size=(width, math.ceil(height / 2)),
        kind=source.kind,
        detail=source.detail,
        scale=scale,
        fit_scale=fit_scale,
        crop_box=crop_box,
        native_pixels=scale >= 0.999999,
    )
    background.close()
    return preview


def _render_image(
    path: Path,
    *,
    max_pixels: int = MAX_PREVIEW_SOURCE_PIXELS,
):
    """Load an image with EXIF orientation and alpha preserved."""
    try:
        from PIL import Image, ImageOps
    except ImportError as exc:
        raise RuntimeError(
            "Image preview requires Pillow. Run: pip install pillow"
        ) from exc

    with Image.open(path) as source:
        frames = int(getattr(source, "n_frames", 1))
        source.seek(0)
        original_size = source.size
        original_pixels = max(1, original_size[0] * original_size[1])
        if max_pixels > 0 and original_pixels > max_pixels:
            reduction = math.sqrt(max_pixels / original_pixels)
            target_size = (
                max(1, int(original_size[0] * reduction)),
                max(1, int(original_size[1] * reduction)),
            )
            try:
                source.draft("RGB", target_size)
            except Exception:
                pass
        else:
            target_size = original_size
        oriented = ImageOps.exif_transpose(source)
        oriented_pixels = max(1, oriented.width * oriented.height)
        if max_pixels > 0 and oriented_pixels > max_pixels:
            reduction = math.sqrt(max_pixels / oriented_pixels)
            oriented_target = (
                max(1, int(oriented.width * reduction)),
                max(1, int(oriented.height * reduction)),
            )
            resampling = getattr(Image, "Resampling", Image)
            oriented.thumbnail(
                oriented_target,
                resample=resampling.LANCZOS,
            )
        has_alpha = (
            "A" in oriented.getbands()
            or "transparency" in getattr(oriented, "info", {})
        )
        image = oriented.convert("RGBA" if has_alpha else "RGB")
        if image.width * image.height > max_pixels > 0:
            resampling = getattr(Image, "Resampling", Image)
            image.thumbnail(target_size, resample=resampling.LANCZOS)
        if image.size == original_size:
            detail = f"{image.width:,} x {image.height:,} px"
        else:
            detail = (
                f"{original_size[0]:,} x {original_size[1]:,} original | "
                f"safe preview {image.width:,} x {image.height:,}"
            )
        if frames > 1:
            detail += f" | frame 1/{frames}"
    return image, "Image", detail


def _render_pdf_with_pymupdf(path: Path):
    """Render the first PDF page with PyMuPDF when it is installed."""
    try:
        import fitz
        from PIL import Image
    except ImportError:
        return None

    document = fitz.open(path)
    try:
        if document.page_count <= 0:
            raise RuntimeError("The PDF contains no pages.")
        page = document.load_page(0)
        page_width = max(1.0, float(page.rect.width))
        scale = min(3.0, max(1.5, 1_600 / page_width))
        pixmap = page.get_pixmap(
            matrix=fitz.Matrix(scale, scale),
            alpha=False,
        )
        image = Image.frombytes(
            "RGB",
            (pixmap.width, pixmap.height),
            pixmap.samples,
        )
        return image, "PDF", f"Page 1 of {document.page_count:,}"
    finally:
        document.close()


def _hidden_startup_info() -> Optional[subprocess.STARTUPINFO]:
    if os.name != "nt":
        return None
    startup = subprocess.STARTUPINFO()
    startup.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    startup.wShowWindow = 0
    return startup


def _render_pdf_with_pdftoppm(path: Path):
    """Render the first PDF page with Poppler when available on PATH."""
    executable = shutil.which("pdftoppm")
    if not executable:
        return None

    try:
        from PIL import Image
    except ImportError as exc:
        raise RuntimeError(
            "PDF preview requires Pillow. Run: pip install pillow"
        ) from exc

    with tempfile.TemporaryDirectory(prefix="mdir_pdf_preview_") as folder:
        output_prefix = Path(folder) / "page"
        poppler_command = [
            executable,
            "-f",
            "1",
            "-l",
            "1",
            "-singlefile",
            "-scale-to-x",
            "1600",
            "-scale-to-y",
            "-1",
            "-png",
            str(path),
            str(output_prefix),
        ]
        command: list[str]
        if os.name == "nt" and Path(executable).suffix.lower() in {
            ".bat",
            ".cmd",
        }:
            command = [
                os.environ.get("COMSPEC", "cmd.exe"),
                "/d",
                "/c",
                subprocess.list2cmdline(poppler_command),
            ]
        else:
            command = poppler_command
        result = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=30,
            startupinfo=_hidden_startup_info(),
            creationflags=(
                subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
            ),
        )
        output_file = output_prefix.with_suffix(".png")
        if result.returncode != 0 or not output_file.exists():
            message = (result.stderr or result.stdout).strip()
            raise RuntimeError(message or "Poppler could not render the PDF.")
        with Image.open(output_file) as source:
            image = source.convert("RGB")
    return image, "PDF", "Page 1"


def _render_pdf(path: Path):
    preview = _render_pdf_with_pymupdf(path)
    if preview is not None:
        return preview
    preview = _render_pdf_with_pdftoppm(path)
    if preview is not None:
        return preview
    raise RuntimeError(
        "PDF preview needs PyMuPDF or Poppler. "
        "Run: pip install pymupdf"
    )


def _column_index(reference: str) -> int:
    """Convert an Excel column reference such as AA to a zero-based index."""
    letters = re.match(r"[A-Za-z]+", reference or "")
    if not letters:
        return 0
    value = 0
    for char in letters.group(0).upper():
        value = value * 26 + ord(char) - ord("A") + 1
    return max(0, value - 1)


def _xlsx_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    values: list[str] = []
    for item in root.findall(".//{*}si"):
        values.append(
            "".join(node.text or "" for node in item.findall(".//{*}t"))
        )
    return values


def _xlsx_first_sheet_path(
    archive: zipfile.ZipFile,
) -> tuple[str, str]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    first_sheet = workbook.find(".//{*}sheet")
    if first_sheet is None:
        raise RuntimeError("The workbook contains no worksheets.")

    sheet_name = first_sheet.attrib.get("name", "Sheet1")
    relation_id = (
        first_sheet.attrib.get(
            "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        )
        or first_sheet.attrib.get("r:id")
    )
    relationships = ET.fromstring(
        archive.read("xl/_rels/workbook.xml.rels")
    )
    target = None
    for relationship in relationships.findall(".//{*}Relationship"):
        if relationship.attrib.get("Id") == relation_id:
            target = relationship.attrib.get("Target")
            break
    if not target:
        raise RuntimeError("The first worksheet relationship is missing.")

    normalized = target.replace("\\", "/").lstrip("/")
    if not normalized.startswith("xl/"):
        normalized = "xl/" + normalized
    return normalized, sheet_name


def _read_xlsx_without_openpyxl(path: Path) -> tuple[list[list[str]], str]:
    """Read a bounded first-sheet sample directly from the XLSX XML package."""
    rows: list[list[str]] = []
    with zipfile.ZipFile(path) as archive:
        shared = _xlsx_shared_strings(archive)
        sheet_path, sheet_name = _xlsx_first_sheet_path(archive)
        with archive.open(sheet_path) as sheet_stream:
            for _, element in ET.iterparse(sheet_stream, events=("end",)):
                if not element.tag.endswith("}row"):
                    continue
                values = [""] * MAX_EXCEL_COLUMNS
                for cell in element.findall("./{*}c"):
                    column = _column_index(cell.attrib.get("r", "A1"))
                    if column >= MAX_EXCEL_COLUMNS:
                        continue
                    cell_type = cell.attrib.get("t", "")
                    value_node = cell.find("./{*}v")
                    value = value_node.text if value_node is not None else ""
                    if cell_type == "s" and value:
                        try:
                            value = shared[int(value)]
                        except (IndexError, ValueError):
                            pass
                    elif cell_type == "inlineStr":
                        value = "".join(
                            node.text or ""
                            for node in cell.findall(".//{*}t")
                        )
                    elif cell_type == "b":
                        value = "TRUE" if value == "1" else "FALSE"
                    values[column] = str(value or "")
                if any(values):
                    rows.append(values)
                element.clear()
                if len(rows) >= MAX_EXCEL_ROWS:
                    break
    return rows, sheet_name


def _read_excel_with_openpyxl(
    path: Path,
) -> Optional[tuple[list[list[str]], str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows: list[list[str]] = []
        for values in sheet.iter_rows(
            min_row=1,
            max_row=MAX_EXCEL_ROWS,
            max_col=MAX_EXCEL_COLUMNS,
            values_only=True,
        ):
            rows.append(
                ["" if value is None else str(value) for value in values]
            )
        return rows, sheet.title
    finally:
        workbook.close()


def _read_legacy_xls(path: Path) -> tuple[list[list[str]], str]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError(
            "Legacy .xls preview requires xlrd. Run: pip install xlrd"
        ) from exc

    workbook = xlrd.open_workbook(path, on_demand=True)
    try:
        sheet = workbook.sheet_by_index(0)
        rows = [
            [
                str(sheet.cell_value(row, column))
                for column in range(
                    min(sheet.ncols, MAX_EXCEL_COLUMNS)
                )
            ]
            for row in range(min(sheet.nrows, MAX_EXCEL_ROWS))
        ]
        return rows, sheet.name
    finally:
        workbook.release_resources()


def _excel_font(size: int, *, bold: bool = False):
    from PIL import ImageFont

    candidates = [
        (
            Path(os.environ.get("WINDIR", r"C:\Windows"))
            / "Fonts"
            / ("malgunbd.ttf" if bold else "malgun.ttf")
        ),
        (
            Path(os.environ.get("WINDIR", r"C:\Windows"))
            / "Fonts"
            / ("arialbd.ttf" if bold else "arial.ttf")
        ),
    ]
    for candidate in candidates:
        if candidate.exists():
            try:
                return ImageFont.truetype(str(candidate), size)
            except OSError:
                pass
    return ImageFont.load_default()


def _draw_excel_grid(
    rows: list[list[str]],
    sheet_name: str,
):
    """Draw a bounded first-sheet sample as an Explorer-style grid."""
    from PIL import Image, ImageDraw

    if not rows:
        rows = [["(empty worksheet)"]]
    column_count = max(1, min(
        MAX_EXCEL_COLUMNS,
        max(len(row) for row in rows),
    ))
    normalized = [
        (row + [""] * column_count)[:column_count]
        for row in rows
    ]

    body_font = _excel_font(19)
    header_font = _excel_font(19, bold=True)
    title_font = _excel_font(22, bold=True)
    measuring = Image.new("RGB", (10, 10), "white")
    measure = ImageDraw.Draw(measuring)

    column_widths: list[int] = []
    for column in range(column_count):
        longest = max(
            [chr(ord("A") + column)]
            + [row[column] for row in normalized],
            key=len,
        )
        bounds = measure.textbbox((0, 0), longest[:40], font=body_font)
        width = max(90, min(300, bounds[2] - bounds[0] + 24))
        column_widths.append(width)

    row_number_width = 58
    title_height = 52
    header_height = 36
    row_height = 34
    width = min(2_000, row_number_width + sum(column_widths) + 2)
    height = title_height + header_height + row_height * len(normalized) + 2
    image = Image.new("RGB", (width, height), "#ffffff")
    draw = ImageDraw.Draw(image)

    draw.rectangle((0, 0, width, title_height), fill="#217346")
    draw.text(
        (16, 11),
        f"Excel preview — {sheet_name}",
        font=title_font,
        fill="white",
    )

    y = title_height
    draw.rectangle((0, y, width, y + header_height), fill="#e6e6e6")
    draw.rectangle(
        (0, y, row_number_width, y + header_height),
        fill="#d5d5d5",
    )
    x = row_number_width
    for column, column_width in enumerate(column_widths):
        if x >= width:
            break
        right = min(width, x + column_width)
        draw.rectangle(
            (x, y, right, y + header_height),
            outline="#a6a6a6",
            fill="#e6e6e6",
        )
        label = ""
        number = column + 1
        while number:
            number, remainder = divmod(number - 1, 26)
            label = chr(ord("A") + remainder) + label
        draw.text(
            (x + 9, y + 6),
            label,
            font=header_font,
            fill="#333333",
        )
        x = right

    for row_index, row in enumerate(normalized, start=1):
        top = title_height + header_height + (row_index - 1) * row_height
        bottom = top + row_height
        draw.rectangle(
            (0, top, row_number_width, bottom),
            outline="#a6a6a6",
            fill="#e6e6e6",
        )
        draw.text(
            (8, top + 5),
            str(row_index),
            font=header_font,
            fill="#444444",
        )
        x = row_number_width
        for column, column_width in enumerate(column_widths):
            if x >= width:
                break
            right = min(width, x + column_width)
            draw.rectangle(
                (x, top, right, bottom),
                outline="#c8c8c8",
                fill="#ffffff",
            )
            value = row[column].replace("\r", " ").replace("\n", " ")
            draw.text(
                (x + 8, top + 5),
                value[:60],
                font=body_font,
                fill="#222222",
            )
            x = right
    return image


def _trim_excel_rows(rows: list[list[str]]) -> list[list[str]]:
    """Remove empty trailing rows and columns from a bounded sheet sample."""
    cleaned = [
        [str(value) for value in row]
        for row in rows
    ]
    while cleaned and not any(value.strip() for value in cleaned[-1]):
        cleaned.pop()
    last_column = 0
    for row in cleaned:
        for index, value in enumerate(row, start=1):
            if value.strip():
                last_column = max(last_column, index)
    if last_column:
        cleaned = [row[:last_column] for row in cleaned]
    return cleaned


def _render_excel(path: Path):
    suffix = path.suffix.lower()
    if suffix == ".xls":
        rows, sheet_name = _read_legacy_xls(path)
    elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        result = _read_excel_with_openpyxl(path)
        if result is None:
            rows, sheet_name = _read_xlsx_without_openpyxl(path)
        else:
            rows, sheet_name = result
    else:
        raise RuntimeError(
            f"{suffix} preview requires an optional Excel reader."
        )

    rows = _trim_excel_rows(rows)
    image = _draw_excel_grid(rows, sheet_name)
    detail = (
        f"Sheet: {sheet_name} | first "
        f"{len(rows):,} rows x "
        f"{max((len(row) for row in rows), default=0):,} columns"
    )
    return image, "Excel", detail


def prepare_document_source(
    path: Path,
    *,
    max_image_pixels: int = MAX_PREVIEW_SOURCE_PIXELS,
) -> DocumentSource:
    """Load one document into a reusable preview source."""
    suffix = path.suffix.lower()
    if suffix in IMAGE_EXTENSIONS:
        image, kind, detail = _render_image(
            path,
            max_pixels=max_image_pixels,
        )
    elif suffix in PDF_EXTENSIONS:
        image, kind, detail = _render_pdf(path)
    elif suffix in EXCEL_EXTENSIONS:
        image, kind, detail = _render_excel(path)
    else:
        raise RuntimeError(f"Preview is not supported for {suffix or 'this file'}.")
    return DocumentSource(image=image, kind=kind, detail=detail)


def render_document_preview(
    path: Path,
    *,
    max_columns: int,
    max_rows: int,
    zoom_factor: float = 1.0,
    native_pixels: bool = False,
    focus: tuple[float, float] = (0.5, 0.5),
) -> DocumentPreview:
    """Render one supported document, reusing the source within the panel."""
    source = prepare_document_source(path)

    try:
        return _render_source_view(
            source,
            max_columns=max_columns,
            max_rows=max_rows,
            zoom_factor=zoom_factor,
            native_pixels=native_pixels,
            focus=focus,
        )
    finally:
        source.close()


class PreviewViewport(ScrollableContainer):
    """Scrollable surface that reserves the mouse wheel for preview zoom."""

    def _zoom_from_wheel(self, event, direction: int) -> None:
        event.stop()
        panel = self.parent
        if panel is not None and hasattr(panel, "handle_wheel_zoom"):
            panel.handle_wheel_zoom(event, direction)

    def on_mouse_scroll_up(self, event: events.MouseScrollUp) -> None:
        self._zoom_from_wheel(event, 1)

    def on_mouse_scroll_down(self, event: events.MouseScrollDown) -> None:
        self._zoom_from_wheel(event, -1)


class DocumentPreviewPanel(Vertical):
    """Non-blocking Explorer-style preview hosted in MDIR's right pane."""

    class CloseRequested(Message):
        """Request restoration of the right file pane."""

    class FullViewRequested(Message):
        """Request the normal F3 full-screen viewer."""

    class OpenRequested(Message):
        """Request opening the document with its Windows application."""

    DEFAULT_CSS = """
    DocumentPreviewPanel {
        width: 100%;
        height: 100%;
        min-height: 0;
        background: $background;
        color: $foreground;
    }

    DocumentPreviewPanel #document_preview_header {
        width: 100%;
        height: 3;
        min-height: 3;
        max-height: 3;
        background: $surface-lighten-1;
        border-bottom: solid $surface-lighten-2;
    }

    DocumentPreviewPanel #document_preview_title {
        width: 1fr;
        height: 3;
        padding: 1;
        color: $foreground;
        text-style: bold;
        text-wrap: nowrap;
        overflow: hidden;
    }

    DocumentPreviewPanel #document_preview_actions {
        width: auto;
        height: 3;
    }

    DocumentPreviewPanel #document_preview_actions Button {
        min-width: 7;
        width: auto;
        height: 3;
        margin: 0 0 0 1;
        border: none;
    }

    DocumentPreviewPanel #document_preview_scroll {
        width: 100%;
        height: 1fr;
        min-height: 4;
        background: $background;
        align: center middle;
        overflow-x: auto;
        overflow-y: auto;
        scrollbar-size: 1 1;
    }

    DocumentPreviewPanel #document_preview_canvas {
        width: auto;
        height: auto;
        min-width: 1;
        min-height: 1;
        background: $background;
        content-align: center middle;
    }

    DocumentPreviewPanel #document_preview_info {
        width: 100%;
        height: 3;
        min-height: 3;
        max-height: 3;
        padding: 0 1;
        background: $panel;
        color: $foreground;
        border-top: solid $surface-lighten-2;
        text-wrap: nowrap;
        overflow: hidden;
    }
    """

    def __init__(self, *, id: Optional[str] = None) -> None:
        super().__init__(id=id)
        self.path: Optional[Path] = None
        self.source: Optional[DocumentSource] = None
        self.preview: Optional[DocumentPreview] = None
        self.render_complete = False
        self.zoom_factor = 1.0
        self.native_pixels = False
        self.focus_point = (0.5, 0.5)
        self._generation = 0
        self._render_timer: Optional[Timer] = None

    def compose(self) -> ComposeResult:
        with Horizontal(id="document_preview_header"):
            yield Static(
                rich_preview_title(),
                id="document_preview_title",
            )
            with Horizontal(id="document_preview_actions"):
                yield Button("Fit", id="document_preview_fit")
                yield Button("1:1", id="document_preview_native")
                yield Button("F3", id="document_preview_full")
                yield Button("Open", id="document_preview_open")
                yield Button("Files", id="document_preview_close")
        with PreviewViewport(id="document_preview_scroll"):
            yield Static(
                "Select an image, PDF, or Excel file in the left pane.",
                id="document_preview_canvas",
            )
        yield Static(
            "Mouse wheel: zoom at pointer | Fit | 1:1 original pixels",
            id="document_preview_info",
        )

    @property
    def canvas(self) -> Static:
        return self.query_one("#document_preview_canvas", Static)

    def _viewport_size(self) -> tuple[int, int]:
        try:
            container = self.query_one(
                "#document_preview_scroll",
                PreviewViewport,
            )
            size = container.content_size
            return max(12, size.width - 2), max(6, size.height - 2)
        except Exception:
            return 80, 24

    def show_path(self, path: Path) -> None:
        """Debounce selection changes before starting document conversion."""
        self.cancel()
        self.path = path
        self.preview = None
        self.render_complete = False
        self.zoom_factor = 1.0
        self.native_pixels = False
        self.focus_point = (0.5, 0.5)
        self.query_one("#document_preview_title", Static).update(
            rich_preview_title(path.name)
        )
        self.canvas.styles.width = "auto"
        self.canvas.styles.height = "auto"
        self.canvas.update(f"Preparing preview...\n\n{path.name}")
        self.query_one("#document_preview_info", Static).update(
            f"{path.suffix.lower().lstrip('.').upper() or 'FILE'} | "
            "Move the left cursor to preview another file"
        )
        self.request_render(delay=0.14)

    def request_render(self, *, delay: float = 0.06) -> None:
        """Schedule a render while retaining the prepared document source."""
        if self.path is None:
            return
        self.render_complete = False
        self._generation += 1
        generation = self._generation
        if self._render_timer is not None:
            self._render_timer.stop()
        self._render_timer = self.set_timer(
            delay,
            lambda: self._start_render(generation),
        )

    def _start_render(self, generation: int) -> None:
        if generation != self._generation or self.path is None:
            return
        columns, rows = self._viewport_size()
        self.canvas.update(
            f"Rendering {self.path.name}...\n\n"
            f"Preview area: {columns} x {rows} cells"
        )
        self._render_document(
            self.path,
            self.source,
            columns,
            rows,
            self.zoom_factor,
            self.native_pixels,
            self.focus_point,
            generation,
        )

    @work(
        thread=True,
        exclusive=True,
        group="document-preview-render",
        exit_on_error=False,
    )
    def _render_document(
        self,
        path: Path,
        source: Optional[DocumentSource],
        columns: int,
        rows: int,
        zoom_factor: float,
        native_pixels: bool,
        focus: tuple[float, float],
        generation: int,
    ) -> None:
        created_source = source is None
        try:
            if source is None:
                source = prepare_document_source(path)
            preview = _render_source_view(
                source,
                max_columns=columns,
                max_rows=rows,
                zoom_factor=zoom_factor,
                native_pixels=native_pixels,
                focus=focus,
            )
        except Exception as exc:
            if created_source and source is not None:
                source.close()
            try:
                self.app.call_from_thread(
                    self._apply_error,
                    str(exc),
                    generation,
                )
            except Exception:
                pass
            return
        try:
            self.app.call_from_thread(
                self._apply_preview,
                preview,
                generation,
                source,
                created_source,
            )
        except Exception:
            if created_source and source is not None:
                source.close()

    def _apply_preview(
        self,
        preview: DocumentPreview,
        generation: int,
        source: DocumentSource,
        created_source: bool,
    ) -> None:
        if generation != self._generation or not self.is_mounted:
            if created_source:
                source.close()
            return
        if created_source:
            previous_source = self.source
            self.source = source
            if previous_source is not None and previous_source is not source:
                previous_source.close()
        self.preview = preview
        self.canvas.styles.width = preview.cell_size[0]
        self.canvas.styles.height = preview.cell_size[1]
        self.canvas.update(preview.content)
        if preview.native_pixels:
            scale_text = "Scale 100% | 1:1 original pixels"
        elif abs(preview.scale - preview.fit_scale) < 1e-6:
            scale_text = f"Fit {preview.scale * 100:.1f}%"
        else:
            scale_text = f"Scale {preview.scale * 100:.1f}%"
        left, top, right, bottom = preview.crop_box
        crop_text = (
            ""
            if preview.crop_box
            == (0, 0, preview.original_size[0], preview.original_size[1])
            else (
                f" | Source area {left:,},{top:,}-"
                f"{right:,},{bottom:,}"
            )
        )
        self.query_one("#document_preview_info", Static).update(
            f"{preview.kind} | {scale_text}{crop_text} | "
            "Wheel: zoom | Fit | 1:1"
        )
        self.render_complete = True

    def _apply_error(self, message: str, generation: int) -> None:
        if generation != self._generation or not self.is_mounted:
            return
        self.preview = None
        self.canvas.styles.width = "auto"
        self.canvas.styles.height = "auto"
        self.canvas.update(
            Text(
                f"Preview unavailable\n\n{message}\n\n"
                "Use F3 or Open to view the original file.",
                style="bold red",
            )
        )
        self.query_one("#document_preview_info", Static).update(
            "Preview failed | F3: full viewer | Open: default Windows app"
        )
        self.render_complete = True

    def cancel(self) -> None:
        self._generation += 1
        if self._render_timer is not None:
            self._render_timer.stop()
            self._render_timer = None
        if self.source is not None:
            self.source.close()
            self.source = None

    def fit_view(self) -> None:
        self.zoom_factor = 1.0
        self.native_pixels = False
        self.focus_point = (0.5, 0.5)
        self.request_render()

    def native_view(self) -> None:
        self.native_pixels = True
        self.request_render()

    def zoom_in(self) -> None:
        if self.native_pixels:
            return
        self.zoom_factor = min(512.0, self.zoom_factor * 1.35)
        self.request_render()

    def zoom_out(self) -> None:
        if self.native_pixels:
            self.native_pixels = False
            if self.preview is not None and self.preview.fit_scale > 0:
                self.zoom_factor = max(
                    1.0,
                    (1.0 / self.preview.fit_scale) / 1.35,
                )
        else:
            self.zoom_factor = max(1.0, self.zoom_factor / 1.35)
        self.request_render()

    def _set_focus_from_mouse(self, event) -> None:
        try:
            container = self.query_one(
                "#document_preview_scroll",
                PreviewViewport,
            )
            region = container.region
            screen_x = getattr(event, "screen_x", event.x)
            screen_y = getattr(event, "screen_y", event.y)
            self.focus_point = (
                max(
                    0.0,
                    min(1.0, (screen_x - region.x) / max(1, region.width)),
                ),
                max(
                    0.0,
                    min(1.0, (screen_y - region.y) / max(1, region.height)),
                ),
            )
        except Exception:
            pass

    def handle_wheel_zoom(self, event, direction: int) -> None:
        """Zoom toward the pointer while keeping the file list focused."""
        self._set_focus_from_mouse(event)
        if direction > 0:
            self.zoom_in()
        else:
            self.zoom_out()

    def on_mouse_scroll_up(self, event: events.MouseScrollUp) -> None:
        event.stop()
        self.handle_wheel_zoom(event, 1)

    def on_mouse_scroll_down(self, event: events.MouseScrollDown) -> None:
        event.stop()
        self.handle_wheel_zoom(event, -1)

    def on_resize(self, event: events.Resize) -> None:
        if (
            self.path is not None
            and self.display
            and self.is_mounted
        ):
            self.request_render(delay=0.10)

    @on(Button.Pressed, "#document_preview_fit")
    def fit_clicked(self, event: Button.Pressed) -> None:
        event.stop()
        self.fit_view()

    @on(Button.Pressed, "#document_preview_native")
    def native_clicked(self, event: Button.Pressed) -> None:
        event.stop()
        self.native_view()

    @on(Button.Pressed, "#document_preview_close")
    def close_clicked(self, event: Button.Pressed) -> None:
        event.stop()
        self.post_message(self.CloseRequested())

    @on(Button.Pressed, "#document_preview_full")
    def full_clicked(self, event: Button.Pressed) -> None:
        event.stop()
        self.post_message(self.FullViewRequested())

    @on(Button.Pressed, "#document_preview_open")
    def open_clicked(self, event: Button.Pressed) -> None:
        event.stop()
        self.post_message(self.OpenRequested())

    def on_unmount(self) -> None:
        self.cancel()
